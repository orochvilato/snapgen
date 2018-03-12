from flask import Response,request,make_response
import json
#from genvisu import use_cache
import datetime


def parse_content(content):
    from lxml import etree
    parser = etree.HTMLParser()
    page   = etree.fromstring(content, parser)
    return page

def maj1l(x):
    return x[0].upper()+x[1:]

def json_response(r):
    resp = Response(json.dumps(r))
    resp.headers['Content-Type'] = 'text/json'
    return resp

from functools import wraps
def cache_function(expires=0):
    def wrap(f):
        @wraps(f)
        def wrapped_f(*args,**kwargs):
            return use_cache(request.url,lambda:f(*args,**kwargs),expires=expires)
        return wrapped_f
    return wrap

def logitem(name,item,fields):
    def wrap(f):
        @wraps(f)
        def wrapped_f(*args,**kwargs):
            log = dict((f,request.args.get(f)) for f in fields if request.args.get(f))
            if not '127.0.0.1' in request.url and not 'api.dev' in request.url:
                log.update({ 'name':name,'item':kwargs.get(item,item),'timestamp':datetime.datetime.now(),'ip':request.environ['REMOTE_ADDR'],'user_agent':request.headers.get('User-Agent')})
                mdbrw.logs.insert_one(log)
            #print args,kwargs
            #print log
            return f(*args,**kwargs)
        return wrapped_f
    return wrap

def getdot(e,k):
    for _k in k.split('.'):
        if _k in e.keys():
            e = e[_k]
        else:
            e = ""
            break
    return e

import unicodedata
def strip_accents(s):
   return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
def normalize(s):
    return strip_accents(s).replace(u'\u2019','').replace('&apos;','').replace(u'\xa0','').encode('utf8').replace(' ','').replace("'",'').replace('-','').replace('\x0a','').replace('\xc5\x93','oe').decode('utf8').lower() if s else s


def dictToXls(data):
    from xlwt import Workbook,XFStyle
    from cStringIO import StringIO
    date_format = XFStyle()
    date_format.num_format_str = 'dd/mm/YYYY'
    stream = StringIO()
    wb = Workbook(encoding='utf8')
    from datetime import datetime,date
    for sheetname in data['sheets']:
        sheet = data['data'][sheetname]
        ws = wb.add_sheet(sheetname)
        for j,field in enumerate(sheet['fields']):
            fieldtxt = field[1] if isinstance(field,tuple) else field
            ws.row(0).write(j,fieldtxt)
        for i,row in enumerate(sheet['data']):
            for j,field in enumerate(sheet['fields']):
                fieldtxt = field[0] if isinstance(field,tuple) else field
                f = row.get(fieldtxt,None)
                if isinstance(f,date) or isinstance(f,datetime):
                    ws.row(i+1).write(j,f,date_format)
                else:
                    ws.row(i+1).write(j,f)

    wb.save(stream)
    return stream.getvalue()

def dictToXlsx(data):
    from openpyxl import Workbook
    from openpyxl.writer.excel import save_virtual_workbook


    wb = Workbook()
    for sheetname in data['sheets']:
        sheet = data['data'][sheetname]
        ws = wb.create_sheet(title=sheetname)
        for j,field in enumerate(sheet['fields']):
            fieldtxt = field[1] if isinstance(field,tuple) else field
            ws.cell(column=j+1,row=1,value=fieldtxt)
        for i,row in enumerate(sheet['data']):
            for j,field in enumerate(sheet['fields']):
                fieldv = field[0] if isinstance(field,tuple) else field
                f = row.get(fieldv,None)
                ws.cell(column=j+1,row=2+i,value=f)

    return save_virtual_workbook(wb)

def xls_response(filename,v):
    import datetime
    output = make_response(v)
    output.headers['Content-Disposition'] = "attachment; filename=%s_%s.xls" % (filename,datetime.datetime.now().strftime('%Y-%m-%d'))
    output.headers['Content-type'] = 'application/vnd.ms-excel'
    return output

def image_response(type,v,filename=None,cookies={},nocache=True):
    r = make_response(v)
    if nocache:
        r.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        r.headers['Pragma'] = 'no-cache'
    r.headers['Content-type'] = "image/%s" % type

    if filename:
        r.headers['Content-Disposition']= "attachment;filename=%s-%s.png" % (filename,datetime.datetime.now().strftime('%Y-%m-%d'))

    for k,v in cookies.iteritems():
        r.set_cookie(k,v)
    #r = Response(v, mimetype="image/%s" % type,headers=headers)

    return r

import pyzmail
from snapgen.config_private import smtp
NOTIFY_ADDRESS = ['observatoireapi@yahoo.com']
SMTP_HOST = smtp['host']

def sendmail(sender,recipients,subject,msg='',attach=[]):
    payload, mail_from, rcpt_to, msg_id=pyzmail.compose_mail(\
        sender, \
        recipients, \
        subject, \
        'utf-8', \
        (msg, 'utf-8'), \
        html=None, \
        attachments=attach)

    #[('attached content', 'text', 'plain', 'text.txt', 'utf-8')]
    smtp_host = SMTP_HOST
    ret=pyzmail.send_mail(payload, mail_from, rcpt_to, smtp_host)

    if isinstance(ret, dict):
        if ret:
            print('failed recipients:', ', '.join(ret.keys()))
    else:
        print('error:', ret)

def api_notify(subject,msg="",attach=[],recipients=NOTIFY_ADDRESS):
    sendmail(('Observatoire API','api@observatoire-democratie.fr'),recipients,subject,msg,attach)
